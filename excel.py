import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

# Create a new workbook
workbook = openpyxl.Workbook()

# Create a new worksheet
ws = workbook.active
ws.title = "Data"

l=2
# Add data to the worksheet
ws.cell(row=1, column=1).value = "Product Name"
ws.cell(row=1, column=2).value = "Price"
ws.cell(row=1, column=3).value = "Size"
ws.cell(row=1, column=4).value = "Rating"
ws.cell(row=1, column=5).value = "Image Link"
ws.cell(row=1, column=6).value = "Product Link"



driver = webdriver.Edge()
driver.maximize_window()

for r in range(1, 5):
    product=driver.get(url="https://www.flipkart.com/search?q=mobiles&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=on&as=off" + str(r))
    for container in driver.find_elements(By.CLASS_NAME, "_2kHMtA"):
        for title in container.find_elements(By.CLASS_NAME, "_4rR01T"):
            product_name = title.get_attribute('title')
            print("Title", title.text)
            ws.cell(row=l, column=1).value = title.text
        for price in container.find_elements(By.CLASS_NAME, "_25b18c"):
                print("Price", price.text)
                ws.cell(row=l, column=2).value = price.text
        for size in container.find_elements(By.CLASS_NAME, "_3Djpdu"):
                print("Size:", size.text)
                ws.cell(row=l, column=3).value = size.text
        for rating in container.find_elements(By.CLASS_NAME, "_1lRcqv"):
                print("Rating: ", rating.text)
                ws.cell(row=l, column=4).value = rating.text
        for hyper_link in container.find_elements(By.CLASS_NAME, "_396cs4"):
                link = hyper_link.get_attribute('src')
                print("Image Link", hyper_link.get_attribute('src'))
                ws.cell(row=l, column=5).value = link
        for hyper in container.find_elements(By.CLASS_NAME, "_1fQZEK"):
                link = hyper.get_attribute('href')
                print("Product Link", link)
                ws.cell(row=l, column=6).value = link
        l += 1

# Save the workbook
workbook.save("Mobiles.xlsx")

# Close the browser window
driver.quit()
